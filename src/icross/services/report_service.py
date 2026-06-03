"""Report generation service using openpyxl.

Registers task handlers for async generation of Excel reports:
  - products: product catalog (name/SKU/price/stock/status)
  - orders: order list (order ID/items/amount/status/date)
  - finance: posting details (product/price/fees/net income)
  - stocks: inventory by warehouse
  - analytics: traffic/conversion/sales data
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Any

from icross.core.storage.ozon_data import ReportStorage, OrderStorage, AnalyticsStorage, WarehouseStorage, ShopStorage
from icross.services.task_queue import register_task

_logger = logging.getLogger(__name__)

_REPORTS_DIR = Path(__file__).resolve().parent.parent.parent.parent / "reports"


def _ensure_dir():
    _REPORTS_DIR.mkdir(parents=True, exist_ok=True)


# ── helpers ──────────────────────────────────────────────────────────────

def _build_workbook(headers: list[str], rows: list[list[Any]], sheet_name: str = "Sheet1"):
    """Build an openpyxl Workbook with a single sheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Write data rows
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center")

    # Auto-width
    for col_idx, header in enumerate(headers, 1):
        max_len = len(str(header))
        for row in rows:
            val = row[col_idx - 1] if col_idx - 1 < len(row) else ""
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[chr(64 + col_idx) if col_idx < 27 else "ZZ"].width = min(max_len + 4, 50)

    return wb


def _save_report(report_id: str, wb) -> tuple[str, int]:
    """Save workbook to disk, return (file_path, file_size)."""
    _ensure_dir()
    path = _REPORTS_DIR / f"{report_id}.xlsx"
    wb.save(str(path))
    file_size = path.stat().st_size
    return str(path), file_size


async def _get_ozon_client():
    """Import and return the Ozon client (lazy import for startup performance)."""
    from icross.services.ozon import get_ozon_client
    return get_ozon_client()


# ── product report ───────────────────────────────────────────────────────

@register_task("report_products")
async def generate_products_report(shop_id: str, report_id: str, **kwargs) -> dict[str, Any]:
    """Generate product catalog Excel report."""
    from icross.core.storage.ozon_data import ProductStorage

    store = ReportStorage()
    await store.update_report(report_id, status="generating")

    try:
        product_store = ProductStorage()
        products = await product_store.list_products(shop_id=shop_id, limit=9999)
        items = products.get("items", products) if isinstance(products, dict) else products

        headers = ["商品名称", "SKU", "Ozon ID", "售价 (₽)", "原价 (₽)", "库存", "状态", "类目"]
        rows = []
        for p in items:
            rows.append([
                p.get("name", p.get("title", "")),
                p.get("sku", p.get("offer_id", "")),
                p.get("product_id", ""),
                p.get("price", p.get("min_price", 0)),
                p.get("old_price", p.get("max_price", 0)),
                p.get("stock", p.get("total_stock", 0)),
                p.get("status", p.get("state", "")),
                p.get("category_name", ""),
            ])

        wb = _build_workbook(headers, rows, "商品列表")
        file_path, file_size = _save_report(report_id, wb)

        await store.update_report(
            report_id, status="completed",
            file_path=file_path, file_size=file_size,
            completed_at=datetime.now().isoformat(),
        )
        return {"success": True, "file_path": file_path, "rows": len(rows)}
    except Exception as e:
        _logger.exception(f"Products report failed: {e}")
        await store.update_report(report_id, status="failed", error=str(e))
        return {"success": False, "error": str(e)}


# ── orders report ────────────────────────────────────────────────────────

@register_task("report_orders")
async def generate_orders_report(shop_id: str, report_id: str, date_from: str = "", date_to: str = "", **kwargs) -> dict[str, Any]:
    """Generate order list Excel report from local storage."""
    store = ReportStorage()
    await store.update_report(report_id, status="generating")

    try:
        order_store = OrderStorage()
        orders = await order_store.list_orders(shop_id=shop_id, limit=9999)
        items = orders if isinstance(orders, list) else orders.get("items", [])

        headers = ["订单号", "商品名称", "数量", "金额 (₽)", "状态", "配送方式", "下单时间"]
        rows = []
        for o in items:
            posting_number = o.get("posting_number", "")
            status = o.get("status", "")
            created_at = o.get("created_at", "")
            delivery = o.get("delivery", o.get("shipping_method", ""))

            products = o.get("products", [])
            if products:
                for pr in products:
                    rows.append([
                        posting_number,
                        pr.get("name", ""),
                        pr.get("quantity", 1),
                        pr.get("price", o.get("total_price", 0)),
                        status,
                        delivery,
                        created_at,
                    ])
            else:
                rows.append([
                    posting_number, "", 1,
                    o.get("total_price", 0),
                    status, delivery, created_at,
                ])

        wb = _build_workbook(headers, rows, "订单列表")
        file_path, file_size = _save_report(report_id, wb)

        await store.update_report(
            report_id, status="completed",
            file_path=file_path, file_size=file_size,
            completed_at=datetime.now().isoformat(),
        )
        return {"success": True, "file_path": file_path, "rows": len(rows)}
    except Exception as e:
        _logger.exception(f"Orders report failed: {e}")
        await store.update_report(report_id, status="failed", error=str(e))
        return {"success": False, "error": str(e)}


# ── finance report ───────────────────────────────────────────────────────

@register_task("report_finance")
async def generate_finance_report(shop_id: str, report_id: str, date_from: str = "", date_to: str = "", **kwargs) -> dict[str, Any]:
    """Generate finance / posting detail Excel report from Ozon API."""
    store = ReportStorage()
    await store.update_report(report_id, status="generating")

    try:
        client = await _get_ozon_client()

        # Try to get realization posting data from API
        rows: list[list[Any]] = []
        headers = ["订单号", "商品", "售价 (₽)", "平台服务费 (₽)", "物流标准费 (₽)",
                    "补贴 (₽)", "扣费合计 (₽)", "净收入 (₽)"]

        if date_from and len(date_from) >= 7:
            year = int(date_from[:4])
            month = int(date_from[5:7])
            result = await client.get_realization_posting(shop_id, month, year)
            postings = result.get("result", []) if isinstance(result, dict) else result
            if isinstance(postings, dict):
                postings = postings.get("rows", postings.get("result", [postings]))

            for p in postings if isinstance(postings, list) else []:
                dc = p.get("delivery_commission", {}) or {}
                rows.append([
                    p.get("posting_number", ""),
                    p.get("product_name", p.get("product", "")),
                    p.get("seller_price_per_instance", 0),
                    dc.get("amount", 0),
                    dc.get("standard_fee", 0),
                    dc.get("bonus", 0),
                    dc.get("total", 0),
                    p.get("payout", p.get("net_income", p.get("seller_reward", 0))),
                ])

        if not rows:
            # Fallback: use local order data
            order_store = OrderStorage()
            orders = await order_store.list_orders(shop_id=shop_id, limit=9999)
            items = orders if isinstance(orders, list) else orders.get("items", [])
            for o in items:
                payout = o.get("payout", o.get("net_income", 0))
                rows.append([
                    o.get("posting_number", ""),
                    o.get("products", [{}])[0].get("name", "") if o.get("products") else "",
                    o.get("total_price", 0),
                    0, 0, 0, 0, payout,
                ])

        wb = _build_workbook(headers, rows, "入账明细")
        file_path, file_size = _save_report(report_id, wb)

        await store.update_report(
            report_id, status="completed",
            file_path=file_path, file_size=file_size,
            completed_at=datetime.now().isoformat(),
        )
        return {"success": True, "file_path": file_path, "rows": len(rows)}
    except Exception as e:
        _logger.exception(f"Finance report failed: {e}")
        await store.update_report(report_id, status="failed", error=str(e))
        return {"success": False, "error": str(e)}


# ── stocks report ────────────────────────────────────────────────────────

@register_task("report_stocks")
async def generate_stocks_report(shop_id: str, report_id: str, **kwargs) -> dict[str, Any]:
    """Generate inventory / stock Excel report."""
    store = ReportStorage()
    await store.update_report(report_id, status="generating")

    try:
        from icross.core.storage.ozon_data import ProductStorage

        product_store = ProductStorage()
        warehouse_store = WarehouseStorage()
        warehouses = await warehouse_store.list_warehouses(shop_id)
        warehouse_map = {w.get("warehouse_id", w.get("id")): w.get("name", w.get("warehouse_name", ""))
                         for w in (warehouses if isinstance(warehouses, list) else warehouses.get("items", []))}

        products = await product_store.list_products(shop_id=shop_id, limit=9999)
        items = products.get("items", products) if isinstance(products, dict) else products

        headers = ["商品名称", "SKU", "仓库", "库存量", "预留量", "可用量", "最低库存预警"]
        rows = []
        for p in items:
            stocks = p.get("stocks", [])
            if stocks:
                for s in stocks:
                    wh_name = warehouse_map.get(s.get("warehouse_id", ""), s.get("warehouse_name", ""))
                    rows.append([
                        p.get("name", p.get("title", "")),
                        p.get("sku", p.get("offer_id", "")),
                        wh_name,
                        s.get("present", s.get("stock", 0)),
                        s.get("reserved", 0),
                        s.get("present", 0) - s.get("reserved", 0),
                        p.get("min_stock", p.get("low_stock_threshold", "")),
                    ])
            else:
                rows.append([
                    p.get("name", p.get("title", "")),
                    p.get("sku", p.get("offer_id", "")),
                    "", p.get("stock", p.get("total_stock", 0)),
                    0, p.get("stock", 0), "",
                ])

        wb = _build_workbook(headers, rows, "库存报表")
        file_path, file_size = _save_report(report_id, wb)

        await store.update_report(
            report_id, status="completed",
            file_path=file_path, file_size=file_size,
            completed_at=datetime.now().isoformat(),
        )
        return {"success": True, "file_path": file_path, "rows": len(rows)}
    except Exception as e:
        _logger.exception(f"Stocks report failed: {e}")
        await store.update_report(report_id, status="failed", error=str(e))
        return {"success": False, "error": str(e)}


# ── analytics report ─────────────────────────────────────────────────────

@register_task("report_analytics")
async def generate_analytics_report(shop_id: str, report_id: str, date_from: str = "", date_to: str = "", **kwargs) -> dict[str, Any]:
    """Generate analytics / traffic data Excel report from Ozon API."""
    store = ReportStorage()
    await store.update_report(report_id, status="generating")

    try:
        client = await _get_ozon_client()
        metrics = ["hits_products_total", "hits_search_total", "session_products_total",
                    "session_search_total", "orders", "ordered_products_cnt", "sales",
                    "sales_with_seller_currency", "ordered_units", "avg_orders_products",
                    "avg_products_sessions", "conversion", "canceled_products_cnt",
                    "canceled_sales", "returns_products_cnt"]
        dimension = ["day", "product"]

        result = await client.get_analytics_data(
            shop_id=shop_id,
            metrics=metrics,
            dimension=dimension,
            date_from=date_from,
            date_to=date_to,
            limit=5000,
        )
        data = result.get("result", []) if isinstance(result, dict) else result

        headers = ["日期", "商品ID", "商品名称", "展示量(搜索)", "展示量(类目)", "访客(搜索)",
                    "访客(类目)", "订单数", "销量", "销售额 (₽)", "退款数", "转化率"]
        rows = []
        for row in data if isinstance(data, list) else []:
            dims = row.get("dimensions", {}) if isinstance(row, dict) else {}
            metrics_data = row.get("metrics", []) if isinstance(row, dict) else []
            rows.append([
                dims.get("day", dims.get("time", "")),
                dims.get("product", dims.get("product_id", "")),
                dims.get("product_name", ""),
                metrics_data[1] if len(metrics_data) > 1 else 0,
                metrics_data[0] if len(metrics_data) > 0 else 0,
                metrics_data[3] if len(metrics_data) > 3 else 0,
                metrics_data[2] if len(metrics_data) > 2 else 0,
                metrics_data[5] if len(metrics_data) > 5 else 0,
                metrics_data[6] if len(metrics_data) > 6 else 0,
                metrics_data[7] if len(metrics_data) > 7 else 0,
                metrics_data[13] if len(metrics_data) > 13 else 0,
                metrics_data[11] if len(metrics_data) > 11 else 0,
            ])

        if not rows:
            # Fallback: use local analytics data
            analytics_store = AnalyticsStorage()
            local = await analytics_store.list_analytics(shop_id=shop_id)
            local_items = local.get("items", []) if isinstance(local, dict) else local
            for a in local_items:
                rows.append([
                    a.get("date", a.get("day", "")),
                    a.get("product_id", ""), a.get("product_name", ""),
                    0, 0, 0, 0,
                    a.get("orders", 0), a.get("ordered_units", 0),
                    a.get("sales", 0), a.get("returns", 0), "",
                ])

        wb = _build_workbook(headers, rows, "数据分析")
        file_path, file_size = _save_report(report_id, wb)

        await store.update_report(
            report_id, status="completed",
            file_path=file_path, file_size=file_size,
            completed_at=datetime.now().isoformat(),
        )
        return {"success": True, "file_path": file_path, "rows": len(rows)}
    except Exception as e:
        _logger.exception(f"Analytics report failed: {e}")
        await store.update_report(report_id, status="failed", error=str(e))
        return {"success": False, "error": str(e)}


# ── scheduled report + send ────────────────────────────────────────


async def generate_and_send_report(
    shop_id: str = "",
    chat_id: str = "",
    platform: str = "feishu",
    report_type: str = "finance",
    **kwargs,
) -> dict:
    """Generate a report and send it to a notification channel.

    Used by the scheduler service for daily report push.

    Args:
        shop_id: Shop ID for the report data.
        chat_id: Target notification chat ID.
        platform: Target platform ("feishu", "telegram", etc.).
        report_type: Report type ("finance", "orders", "products", "stocks", "analytics").
        **kwargs: Additional report parameters (date_from, date_to, etc.).

    Returns:
        Dict with send result.
    """
    _logger.info("Generating and sending %s report for shop %s", report_type, shop_id)

    # 1. Generate the report
    report_id = datetime.now().strftime("%Y%m%d_%H%M%S") + "_" + shop_id[:4]
    task_map = {
        "finance": generate_finance_report,
        "orders": generate_orders_report,
        "products": generate_products_report,
        "stocks": generate_stocks_report,
        "analytics": generate_analytics_report,
    }

    handler = task_map.get(report_type)
    if not handler:
        return {"success": False, "error": f"Unknown report type: {report_type}"}

    result = await handler(shop_id=shop_id, report_id=report_id, **kwargs)

    if not result.get("success"):
        _logger.warning("Report generation failed: %s", result.get("error"))
        return result

    # 2. Send notification with summary
    title = f"每日{report_type}报告"
    content = f"报告已生成: {result.get('file_path', '')}\n数据行数: {result.get('rows', 0)}"

    try:
        from icross.services.notification import get_notification_service

        ns = get_notification_service()
        send_result = await ns.send(
            title=title,
            content=content,
            level="info",
            chat_id=chat_id or None,
            platform=platform,
        )
        _logger.info("Report notification sent: %s", send_result)
        return {"success": True, "report": result, "notification": send_result}
    except Exception as e:
        _logger.warning("Failed to send report notification: %s", e)
        return {"success": True, "report": result, "notification_error": str(e)}


# Register with scheduler service
try:
    from icross.services.scheduler import register_job_handler

    register_job_handler("daily_sales_report", "icross.services.report_service.generate_and_send_report")
except ImportError:
    pass  # scheduler module not yet available
